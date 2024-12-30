# contentscraper.py

import os
from typing import Dict, Any, List
import logging
import google.generativeai as genai
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from dotenv import load_dotenv
import base64
from PIL import Image
import io
import datetime
import json
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from rich import print as rprint
from rich.traceback import install
from loguru import logger

# Install rich traceback handler
install(show_locals=True)

# Setup loguru
logger.add("debug.log", format="{time} {level} {message}", level="DEBUG", rotation="10 MB")

# Add Pydantic models
class SpreadsheetAnalysis(BaseModel):
    summary: str
    key_points: List[str]
    insights: List[str]

class AnalysisResponse(BaseModel):
    type: str
    metadata: dict
    analysis: SpreadsheetAnalysis

class ContentScraper:
    def __init__(self, initialize_models=True):
        load_dotenv()
        
        # Setup logging
        self.logger = logging.getLogger(__name__)
        
        if initialize_models:
            # Only initialize models if flag is True
            GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY")
            if not GOOGLE_API_KEY:
                raise ValueError("GOOGLE_API_KEY not found in environment variables")
                
            genai.configure(api_key=GOOGLE_API_KEY)
        
        # Initialize models (will use existing configuration if already done)
        self.text_model = genai.GenerativeModel('gemini-1.5-pro')
        self.vision_model = genai.GenerativeModel('gemini-1.5-flash')
        
        # Removed the setup_google_auth() call since we're not using Google Docs

    def process_input(self, file_path: str) -> Dict:
        """Process input based on file type"""
        file_ext = os.path.splitext(file_path)[1].lower()
        
        try:
            if file_ext in ['.png', '.jpg', '.jpeg']:
                return self._process_with_vision(file_path)
            elif file_ext in ['.xlsx', '.xls']:
                return self._process_spreadsheet(file_path)
            elif file_ext in ['.doc', '.docx']:
                return self._process_document(file_path)
            else:
                return {
                    "type": "unsupported",
                    "message": f"File type {file_ext} not supported"
                }
        except Exception as e:
            self.logger.error(f"Processing failed: {str(e)}")
            return {
                "type": "error",
                "error": str(e)
            }

    def _process_text(self, file_path: str) -> Dict[str, Any]:
        """Process text using Gemini Pro"""
        try:
            with open(file_path, 'r') as f:
                content = f.read()

            # Use Gemini for content analysis
            prompt = """
            Analyze this content and provide:
            1. A brief summary
            2. Key topics or themes
            3. Important metadata
            4. Suggested tags
            Format the response as JSON.
            
            Content: {content}
            """
            
            response = self.text_model.generate_content(prompt.format(content=content))
            analysis = response.text  # This will be JSON formatted

            return {
                "type": "text",
                "original_content": content,
                "analysis": analysis,
                "metadata": self._extract_basic_metadata(file_path)
            }
        except Exception as e:
            self.logger.error(f"Text processing failed: {str(e)}")
            raise

    def _process_with_vision(self, file_path: str) -> Dict[str, Any]:
        """Process images/unknown files with Gemini Vision"""
        try:
            # Load image
            image = Image.open(file_path)
            
            # Convert image to RGB if necessary
            if image.mode != 'RGB':
                image = image.convert('RGB')
            
            # Enhanced retail-focused prompt
            prompt = """Analyze this image in the context of retail and fashion business.
            Provide a structured analysis with these exact keys (no markdown or code blocks):
            {
                "retail_elements": {
                    "store_layout": "",
                    "product_displays": "",
                    "branding_elements": "",
                    "customer_experience": ""
                },
                "product_analysis": {
                    "visible_products": "",
                    "presentation": "",
                    "quality_indicators": "",
                    "price_positioning": ""
                },
                "visual_merchandising": {
                    "display_techniques": "",
                    "color_schemes": "",
                    "lighting": "",
                    "space_usage": ""
                },
                "branding": {
                    "logo": "",
                    "identity": "",
                    "signage": "",
                    "positioning": ""
                },
                "business_insights": {
                    "target_market": "",
                    "advantages": "",
                    "customer_flow": "",
                    "sales_strategy": ""
                }
            }"""
            
            try:
                response = self.vision_model.generate_content(
                    contents=[prompt, image],
                    generation_config={
                        "temperature": 0.4,
                        "top_p": 0.8,
                        "top_k": 40
                    }
                )
                
                # Clean and parse the response
                analysis_text = response.text.strip()
                if analysis_text.startswith('```'):
                    analysis_text = analysis_text.split('```')[1]
                    if analysis_text.startswith('json'):
                        analysis_text = analysis_text[4:]
                
                try:
                    structured_analysis = json.loads(analysis_text)
                except json.JSONDecodeError:
                    structured_analysis = self._extract_section_structured(analysis_text)
                
                return {
                    "type": "image",
                    "analysis": structured_analysis,
                    "metadata": self._extract_basic_metadata(file_path),
                    "file_info": {
                        "name": os.path.basename(file_path),
                        "directory": os.path.dirname(file_path),
                        "analysis_timestamp": datetime.datetime.now().isoformat()
                    }
                }
                
            except Exception as e:
                self.logger.error(f"Vision model error: {str(e)}")
                return {
                    "type": "image",
                    "analysis": f"Vision analysis unavailable: {str(e)}",
                    "metadata": self._extract_basic_metadata(file_path)
                }
                
        except Exception as e:
            self.logger.error(f"Vision processing failed: {str(e)}")
            raise

    def _extract_section(self, text: str, section_name: str) -> Dict[str, Any]:
        """Helper method to extract and structure sections from the analysis text"""
        try:
            # Find the section in the text
            start = text.find(section_name)
            if start == -1:
                return {}
            
            # Find the next section or end of text
            next_section_start = float('inf')
            for section in ["RETAIL ELEMENTS", "PRODUCT ANALYSIS", "VISUAL MERCHANDISING", "BRANDING", "BUSINESS INSIGHTS"]:
                if section != section_name:
                    pos = text.find(section, start + len(section_name))
                    if pos != -1 and pos < next_section_start:
                        next_section_start = pos
            
            # Extract the section content
            section_text = text[start + len(section_name):next_section_start].strip()
            
            # Convert bullet points to dictionary items
            items = {}
            for line in section_text.split('\n'):
                line = line.strip()
                if line.startswith('-'):
                    key_value = line[1:].strip().split(':', 1)
                    if len(key_value) == 2:
                        items[key_value[0].strip()] = key_value[1].strip()
                    else:
                        items[f"item_{len(items)+1}"] = key_value[0].strip()
                    
            return items
        except Exception:
            return {"raw": text}

    def _process_audio(self, file_path: str) -> Dict[str, Any]:
        """Process audio - first transcribe, then analyze with Gemini"""
        try:
            # First get transcription (you might want to use Whisper or other service)
            # For now, using placeholder
            transcript = "Placeholder transcript"
            
            # Analyze transcript with Gemini
            prompt = """
            Analyze this audio transcript and provide:
            1. Summary of the conversation/content
            2. Key points discussed
            3. Speakers identified (if any)
            4. Important timestamps/segments
            Format the response as JSON.
            
            Transcript: {transcript}
            """
            
            response = self.text_model.generate_content(prompt.format(transcript=transcript))
            analysis = response.text

            return {
                "type": "audio",
                "transcript": transcript,
                "analysis": analysis,
                "metadata": self._extract_basic_metadata(file_path)
            }
        except Exception as e:
            self.logger.error(f"Audio processing failed: {str(e)}")
            raise

    def _process_video(self, file_path: str) -> Dict[str, Any]:
        """Process video - extract frames and analyze with Gemini Vision"""
        try:
            # Extract key frames (implementation needed)
            frames = ["placeholder_frame"]
            
            # Analyze frames with Gemini Vision
            analyses = []
            for frame in frames:
                response = self.vision_model.generate_content([
                    "Describe this video frame in detail, including any actions, objects, or text visible.",
                    frame
                ])
                analyses.append(response.text)

            return {
                "type": "video",
                "frame_analyses": analyses,
                "metadata": self._extract_basic_metadata(file_path)
            }
        except Exception as e:
            self.logger.error(f"Video processing failed: {str(e)}")
            raise

    def _determine_input_type(self, file_path: str) -> str:
        """Determine the type of input file"""
        ext = os.path.splitext(file_path)[1].lower()
        if ext in ['.jpg', '.jpeg', '.png']:
            return 'image'
        elif ext in ['.txt', '.doc', '.docx', '.pdf']:
            return 'text'
        elif ext in ['.mp3', '.wav']:
            return 'audio'
        elif ext in ['.mp4', '.mov']:
            return 'video'
        return 'unknown'

    def _extract_basic_metadata(self, file_path: str) -> Dict[str, Any]:
        """Extract basic file metadata"""
        stat = os.stat(file_path)
        return {
            "size": stat.st_size,
            "created": stat.st_ctime,
            "modified": stat.st_mtime,
            "file_type": os.path.splitext(file_path)[1]
        }

    def _print_analysis(self, result: Dict[str, Any]) -> None:
        """Pretty print the analysis results"""
        import json
        
        print("\nAnalysis Result:")
        print("=" * 80)
        
        # Print file info
        print(f"\nFile: {result['file_info']['name']}")
        print(f"Directory: {result['file_info']['directory']}")
        print(f"Analyzed at: {result['file_info']['analysis_timestamp']}")
        
        # Print analysis sections
        print("\nAnalysis:")
        print("-" * 80)
        for section, content in result['analysis'].items():
            if section != 'raw_analysis':
                print(f"\n{section.upper()}:")
                for key, value in content.items():
                    print(f"  {key}: {value}")
        
        # Print metadata
        print("\nMetadata:")
        print("-" * 80)
        for key, value in result['metadata'].items():
            print(f"  {key}: {value}")

    def _process_spreadsheet(self, file_path: str) -> Dict:
        """Process Excel files"""
        try:
            # Verify file exists
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File not found: {file_path}")
            
            # Log file details
            logger.info(f"Processing Excel file: {file_path}")
            file_size = os.path.getsize(file_path)
            logger.info(f"File size: {file_size} bytes")

            # Load workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            logger.info(f"Sheets found: {wb.sheetnames}")

            # Collect content
            content = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                logger.debug(f"Reading sheet: {sheet}")
                rows = list(ws.rows)[:6]
                logger.debug(f"Found {len(rows)} rows")
                
                for row in rows:
                    row_data = ",".join(str(cell.value or '') for cell in row)
                    content.append(row_data)
                    logger.debug(f"Row data: {row_data}")

            # Get analysis
            logger.info("Sending to Gemini for analysis")
            prompt = f"""Analyze this Excel spreadsheet content and provide insights in this exact format (no additional text, just the JSON):
{{
    "summary": "A clear summary of the spreadsheet content",
    "key_points": [
        "Key point 1",
        "Key point 2",
        "Key point 3"
    ],
    "insights": [
        "Business insight 1",
        "Business insight 2",
        "Business insight 3"
    ]
}}

Content to analyze:
{chr(10).join(content)}"""

            response = self.text_model.generate_content(prompt)
            logger.debug(f"Raw response: {response.text}")

            # Parse response
            try:
                cleaned = self._clean_json_response(response.text)
                logger.debug(f"Cleaned JSON: {cleaned}")
                
                analysis = SpreadsheetAnalysis.model_validate_json(cleaned)
                logger.info("Successfully parsed response")
                
                return {
                    "type": "spreadsheet",
                    "metadata": self._extract_basic_metadata(file_path),
                    "analysis": analysis.model_dump()
                }
                
            except Exception as e:
                logger.exception("Failed to parse response")
                raise
                
        except Exception as e:
            logger.exception("Excel processing failed")
            rprint("[red]Processing Error:[/red]", e)
            return {
                "type": "error",
                "metadata": self._extract_basic_metadata(file_path),
                "analysis": {
                    "type": "error",
                    "error": str(e)
                }
            }

    def _clean_json_response(self, text: str) -> str:
        """Clean and extract JSON from response"""
        text = text.strip()
        
        # Remove markdown code blocks if present
        if text.startswith('```') and text.endswith('```'):
            text = text[3:-3]
        if text.startswith('```json'):
            text = text[7:]
        
        # Find JSON boundaries
        start = text.find('{')
        end = text.rfind('}') + 1
        
        if start >= 0 and end > start:
            text = text[start:end]
        
        return text

    def _validate_json_structure(self, json_str: str) -> Dict:
        """Validate JSON has required structure"""
        required_keys = {
            "purpose": str,
            "key_findings": {
                "data_points": list,
                "patterns": list,
                "trends": list
            },
            "business_insights": {
                "implications": list,
                "opportunities": list,
                "risks": list
            },
            "recommendations": list
        }

        try:
            data = json.loads(json_str)
            # Validate structure (implement validation logic)
            return data
        except json.JSONDecodeError:
            self.logger.error("Invalid JSON structure")
            raise

if __name__ == "__main__":
    scraper = ContentScraper()
    
    print("Testing ContentScraper...")
    print("\nAvailable image files in assets directory:")
    
    assets_dir = "assets"
    for file in os.listdir(assets_dir):
        if file.endswith(('.jpg', '.jpeg', '.png')):
            print(f"- {os.path.join(assets_dir, file)}")
    
    default_file = "assets/Jakkimulls Store.png"
    test_file = input(f"\nEnter path to test file (press Enter for default: {default_file}): ").strip()
    
    if not test_file:
        test_file = default_file
    
    if test_file:
        try:
            print(f"\nAnalyzing: {test_file}")
            result = scraper.process_input(test_file)
            scraper._print_analysis(result)
        except Exception as e:
            print(f"\nError during processing: {e}")